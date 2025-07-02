import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# This code reads a html file as well as a clean Excel file. It then outputs a new Excel file with all the extracted
# information from the html.


# Function to extract data from HTML
def extract_reviews_from_html(html_file):
    # Read the HTML content from the file
    with open(html_file, 'r', encoding='utf-8') as file:
        html_content = file.read()

    soup = BeautifulSoup(html_content, 'html.parser')

    reviews = soup.find_all('div', class_='c-siteReview g-bg-gray10 u-grid g-outer-spacing-bottom-large')

    data = []
    for review in reviews:
        # For each review, we extract the desired information
        score = review.find('div', class_='c-siteReviewScore').find('span').text.strip()
        username = review.find('a', class_='c-siteReviewHeader_username').text.strip()
        date = review.find('div', class_='c-siteReviewHeader_reviewDate').text.strip()
        review_text = review.find('div', class_='c-siteReview_quote').find('span').text.strip()
        platform = review.find('div', class_='c-siteReview_platform g-text-bold g-color-gray80 g-text-xsmall u-text-'
                                             'right u-text-uppercase').text.strip()

        # We then add the extracted information to the output excel
        data.append({
            'Date_of_Review': date,
            'Gaming_Platform': platform,
            'User': username,
            'Reviews': review_text,
            'User_Score': score,
        })

    return pd.DataFrame(data)


def append_data_to_excel(input_file_path, output_file_path, new_data_df):
    # Load existing workbook
    wb = load_workbook(input_file_path)
    ws = wb.active

    # Find the last column
    last_col = ws.max_column

    # Write the new data to the existing workbook
    for r_idx, row in enumerate(dataframe_to_rows(new_data_df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=last_col + 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Copy the formatting from the first row to the new columns
    for col in range(last_col + 1, last_col + 1 + len(new_data_df.columns)):
        col_letter = ws.cell(row=1, column=col).column_letter
        ws.column_dimensions[col_letter].width = ws.column_dimensions[ws.cell(row=1, column=1).column_letter].width

    wb.save(output_file_path)
    print(f"New columns added and saved to {output_file_path}")


html_file = 'eldenring_ps5.html'
input_excel_file = 'eldenring_temp.xlsx'
output_excel_file = 'eldenring_reviews_ps5.xlsx'

new_data_df = extract_reviews_from_html(html_file)

append_data_to_excel(input_excel_file, output_excel_file, new_data_df)