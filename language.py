import pandas as pd
from langdetect import detect
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

# Load the Excel file
input_file = 'eldenring_reviews.xlsx'
output_file = 'eldenring_reviews_sorted.xlsx'
df = pd.read_excel(input_file)


def detect_language(review):
    try:
        return detect(review)
    except:
        return 'unknown'


df['Language'] = df['Reviews'].apply(detect_language)

df_sorted = df.sort_values(by='Language')

# Create a new workbook and add a worksheet
wb = Workbook()
ws = wb.active
ws.title = 'Sorted Reviews'

# Append the DataFrame to the worksheet
for r_idx, row in enumerate(dataframe_to_rows(df_sorted, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        # Ensure the header is not bold
        if r_idx == 1:
            cell.font = Font(bold=False)

# Ensure date format is preserved
for cell in ws[1]:
    if cell.value in df_sorted.columns and 'date' in cell.value.lower():
        col_letter = cell.column_letter
        for row in ws.iter_rows(min_row=2, min_col=cell.column, max_col=cell.column):
            for c in row:
                if isinstance(c.value, pd.Timestamp):
                    c.number_format = 'MM/DD/YYYY'

# Adjust column widths
max_width = 50  # Set a maximum width for the review column
for column_cells in ws.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    adjusted_width = min(max_width, length)
    ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

wb.save(output_file)

print(f"Processed file saved as {output_file}")
