import pandas as pd
from textblob import TextBlob
from openpyxl import load_workbook


# Function to load data from Excel
def load_data_from_excel(file_path):
    df = pd.read_excel(file_path, engine='openpyxl')
    return df


# Function to preprocess reviews
def preprocess_review(review, neutral_phrases):
    for phrase in neutral_phrases:
        review = review.replace(phrase, "NEUTRAL_PHRASE")
    return review


# Function to perform sentiment analysis using TextBlob
def perform_sentiment_analysis(input_file_path, output_file_path, neutral_phrases):
    data = load_data_from_excel(input_file_path)

    sentiment_scores = []
    for review in data['Reviews']:
        preprocessed_review = preprocess_review(review, neutral_phrases)
        blob = TextBlob(preprocessed_review)
        sentiment_score = blob.sentiment.polarity  # Get sentiment polarity score (-1 to 1)
        sentiment_scores.append(sentiment_score)

    data['Sentiment_Score'] = sentiment_scores

    # Load the original workbook
    wb = load_workbook(input_file_path)
    ws = wb.active

    # Find the first empty column
    max_column = ws.max_column
    sentiment_col_idx = max_column + 1

    # Write the Sentiment_Score header
    ws.cell(row=1, column=sentiment_col_idx, value='Textblob_Sentiment_Score')

    # Write the updated DataFrame to the existing workbook
    for row_idx, score in enumerate(sentiment_scores, start=2):
        ws.cell(row=row_idx, column=sentiment_col_idx, value=score)

    # Save the workbook to the new file
    wb.save(output_file_path)
    print(f"Sentiment analysis results saved to {output_file_path}")


# Input file can be altered depending on the game
input_file_path = 'part5.xlsx'
output_file_path = 'part5_textblob_sentiment.xlsx' # This file's name can be changed according to the input file
neutral_phrases = ["Dark Souls", "Dark souls", "dark Souls", "dark souls", "DARK SOULS", "Demons", "Bloodborne", "Sekiro"]

perform_sentiment_analysis(input_file_path, output_file_path, neutral_phrases)