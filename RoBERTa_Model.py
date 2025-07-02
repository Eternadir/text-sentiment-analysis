import pandas as pd
from transformers import AutoTokenizer, AutoModelForSequenceClassification
import numpy as np
import warnings
from openpyxl import load_workbook

# Ignore some warnings that aren't a concern at the moment
warnings.filterwarnings('ignore', category=FutureWarning, module='huggingface_hub.file_download')
warnings.filterwarnings('ignore', category=UserWarning, module='transformers.modeling_utils')

# Function to preprocess reviews
def preprocess_review(review, neutral_phrases):
    for phrase in neutral_phrases:
        review = review.replace(phrase, "NEUTRAL_PHRASE")
    return review

# Input file can be altered depending on the game
input_file_path = 'part5.xlsx'
df = pd.read_excel(input_file_path, engine='openpyxl')

# Ensure the reviews column exists
if 'Reviews' not in df.columns:
    raise ValueError("The Excel file must contain a 'Reviews' column.")

# Load the tokenizer and model
model_name = 'cardiffnlp/twitter-roberta-base-sentiment-latest'
tokenizer = AutoTokenizer.from_pretrained(model_name)
model = AutoModelForSequenceClassification.from_pretrained(model_name)

# Function to get sentiment score
def get_sentiment_score(review):
    inputs = tokenizer(review, return_tensors="pt", truncation=True, padding=True, max_length=512)
    outputs = model(**inputs)
    scores = outputs.logits.detach().numpy()[0]
    # Convert logits to probabilities
    scores = np.exp(scores) / np.exp(scores).sum()
    # Calculate sentiment score
    sentiment_score = scores[2] - scores[0]  # positive - negative
    return sentiment_score

neutral_phrases = ["Dark Souls", "Dark souls", "dark Souls", "dark souls", "DARK SOULS"]
df['Reviews'] = df['Reviews'].apply(preprocess_review, neutral_phrases=neutral_phrases)
df['Sentiment_Score'] = df['Reviews'].apply(get_sentiment_score)

# Load the original workbook
wb = load_workbook(input_file_path)
ws = wb.active

# Find the first empty column
max_column = ws.max_column
sentiment_col_idx = max_column + 1

# Write the Sentiment_Score header
ws.cell(row=1, column=sentiment_col_idx, value='RoBERTa_Sentiment_Score')

# Write the updated DataFrame to the existing workbook
for row_idx, score in enumerate(df['Sentiment_Score'], start=2):
    ws.cell(row=row_idx, column=sentiment_col_idx, value=score)

# Save the workbook to the new file
output_file_path = 'part5_roberta_sentiment.xlsx' # This file's name can be changed according to the input file
wb.save(output_file_path)

print(f"Sentiment analysis complete. Results saved to {output_file_path}")